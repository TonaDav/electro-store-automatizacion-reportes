from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def generar_reporte(inventario):
    # Creamos el workbook
    mi_workbook = Workbook()

    # Seleccionamos y renombramos la hoja de cáculo creada automáticamente
    hoja_resumen = mi_workbook.active
    hoja_resumen.title = "Resumen"

    # Creamos las demás hojas de cálculo
    hoja_detalle_ventas = mi_workbook.create_sheet("Detalle de Ventas")
    hoja_alertas_stock = mi_workbook.create_sheet("Alertas de Stock")

    # Hoja "Resumen"
    hoja_resumen.column_dimensions["A"].width = 25
    hoja_resumen.column_dimensions["B"].width = 30

    hoja_resumen["A1"] = "Ingresos Totales"
    hoja_resumen["B1"] = inventario.ingresos_totales()
    hoja_resumen["B1"].number_format = '"S/" #,##0.00'
    hoja_resumen["A1"].font = Font(bold=True, size=11)
    hoja_resumen["A1"].fill = PatternFill(start_color="00B0F0", fill_type="solid")

    hoja_resumen["A3"] = "Producto más vendido"
    hoja_resumen["B3"] = inventario.producto_mas_vendido()
    hoja_resumen["A3"].font = Font(bold=True, size=11)
    hoja_resumen["A3"].fill = PatternFill(start_color="92D050", fill_type="solid")

    hoja_resumen["A5"] = "Mejor Vendedor"
    hoja_resumen["B5"] = inventario.mejor_vendedor()
    hoja_resumen["A5"].font = Font(bold=True, size=11)
    hoja_resumen["A5"].fill = PatternFill(start_color="FFC000", fill_type="solid")

    hoja_resumen["A7"] = "Cliente más Frecuente"
    hoja_resumen["B7"] = inventario.cliente_mas_frecuente()
    hoja_resumen["A7"].font = Font(bold=True, size=11)
    hoja_resumen["A7"].fill = PatternFill(start_color="FF0000", fill_type="solid")

    # Hoja "Detalle de Ventas"
    hoja_detalle_ventas.append(["# Venta", "ID Producto", "ID Cliente", "ID Vendedor", "Cantidad", "Fecha", "Total"])
    columnas = ["A", "B", "C", "D", "E", "F", "G"]

    # Aumentamos el ancho de las columnas
    for columna in columnas:
        hoja_detalle_ventas.column_dimensions[columna].width = 25

    # Creamos los encabezados de nuestra tabla
    for columna in columnas:
        hoja_detalle_ventas[columna + "1"].font = Font(bold=True, size=12, color="FFFFFF")
        hoja_detalle_ventas[columna + "1"].fill = PatternFill(start_color="FFC000", fill_type="solid")

    # Llenamos la tabla
    for venta in inventario.get_ventas():
        hoja_detalle_ventas.append(
            [venta.id_venta, venta.producto.id_producto, venta.cliente.id_cliente, venta.vendedor.id_vendedor,
             venta.cantidad, venta.fecha, venta.calcular_total()])

    for fila in range(2, len(inventario.get_ventas()) + 2):
        hoja_detalle_ventas[f"G{fila}"].number_format = '"S/" #,##0.00'

    # Hoja "Alertas de Stock"
    hoja_alertas_stock["A1"] = "ID Producto"
    hoja_alertas_stock["B1"] = "Nombre"
    hoja_alertas_stock["C1"] = "Stock"

    columnas = ["A", "B", "C"]

    for columna in columnas:
        hoja_alertas_stock[columna + "1"].font = Font(bold=True, size=12)
        hoja_alertas_stock[columna + "1"].fill = PatternFill(start_color="92D050", fill_type="solid")

    hoja_alertas_stock.column_dimensions["A"].width = 15
    hoja_alertas_stock.column_dimensions["B"].width = 35
    hoja_alertas_stock.column_dimensions["A"].width = 25

    for producto in inventario.productos_bajos_stock():
        hoja_alertas_stock.append([producto.id_producto, producto.nombre, producto.stock])

    for fila in range(2, len(inventario.productos_bajos_stock()) + 2):
        celda = hoja_alertas_stock[f"C{fila}"]

        if celda.value > 0:
            celda.fill = PatternFill(start_color="FFFF00", fill_type="solid")
        else:
            celda.fill = PatternFill(start_color="FF0000", fill_type="solid")

    mi_workbook.save("reporte_electrostore.xlsx")

if __name__ == "__main__":
    from generador_datos import generar_datos
    inventario = generar_datos()
    generar_reporte(inventario)